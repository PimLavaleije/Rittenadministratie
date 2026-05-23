import os
import io
from datetime import date, datetime
from decimal import Decimal

from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, jsonify, send_file, abort,
)
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import func, extract

from models import db, Vehicle, Driver, Trip
from odoo_connector import OdooConnector

load_dotenv()

app = Flask(__name__)
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "dev-change-this-in-production")
import ssl as _ssl
_db_url = os.getenv("DATABASE_URL", "sqlite:///ritten.db")
if _db_url.startswith("postgresql://"):
    _db_url = _db_url.replace("postgresql://", "postgresql+pg8000://", 1).replace("?sslmode=require", "")
    app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {"connect_args": {"ssl_context": _ssl.create_default_context()}}
app.config["SQLALCHEMY_DATABASE_URI"] = _db_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["UPLOAD_FOLDER"] = "uploads"
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # 10 MB max upload

db.init_app(app)
odoo = OdooConnector()


@app.before_request
def create_tables():
    db.create_all()


# ── Dashboard ──────────────────────────────────────────────────────────────

@app.route("/")
def dashboard():
    jaar = request.args.get("jaar", date.today().year, type=int)
    maand = request.args.get("maand", date.today().month, type=int)

    base = Trip.query.filter(extract("year", Trip.datum) == jaar)
    if maand:
        base = base.filter(extract("month", Trip.datum) == maand)

    totaal_zakelijk = sum(
        t.kilometers for t in base.filter_by(type="zakelijk").all()
    )
    totaal_prive = sum(
        t.kilometers for t in base.filter_by(type="prive").all()
    )

    # Kilometers per voertuig
    per_voertuig = (
        db.session.query(Vehicle.kenteken, func.sum(Trip.eindstand_km - Trip.beginstand_km))
        .join(Trip)
        .filter(extract("year", Trip.datum) == jaar)
        .group_by(Vehicle.kenteken)
        .all()
    )

    # Kilometers per maand (huidig jaar)
    per_maand = (
        db.session.query(
            extract("month", Trip.datum).label("maand"),
            func.sum(Trip.eindstand_km - Trip.beginstand_km).label("km"),
        )
        .filter(extract("year", Trip.datum) == jaar)
        .group_by("maand")
        .order_by("maand")
        .all()
    )
    maand_labels = ["Jan", "Feb", "Mrt", "Apr", "Mei", "Jun", "Jul", "Aug", "Sep", "Okt", "Nov", "Dec"]
    maand_data = [0.0] * 12
    for rij in per_maand:
        maand_data[int(rij.maand) - 1] = float(rij.km or 0)

    recente_ritten = (
        Trip.query.filter(extract("year", Trip.datum) == jaar)
        .order_by(Trip.datum.desc(), Trip.aangemaakt_op.desc())
        .limit(10).all()
    )

    jaren = [r[0] for r in db.session.query(extract("year", Trip.datum)).distinct().order_by(extract("year", Trip.datum).desc()).all()]
    if jaar not in jaren:
        jaren.insert(0, jaar)

    return render_template(
        "dashboard.html",
        totaal_zakelijk=totaal_zakelijk,
        totaal_prive=totaal_prive,
        per_voertuig=per_voertuig,
        maand_labels=maand_labels,
        maand_data=maand_data,
        recente_ritten=recente_ritten,
        jaar=jaar,
        maand=maand,
        jaren=jaren,
        maand_namen=maand_labels,
    )


# ── Rittenlijst ────────────────────────────────────────────────────────────

@app.route("/ritten")
def ritten_lijst():
    zoek = request.args.get("zoek", "")
    type_filter = request.args.get("type", "")
    voertuig_filter = request.args.get("voertuig", "", type=int)
    jaar_filter = request.args.get("jaar", 0, type=int)
    maand_filter = request.args.get("maand", 0, type=int)
    pagina = request.args.get("pagina", 1, type=int)

    query = Trip.query.order_by(Trip.datum.desc(), Trip.aangemaakt_op.desc())

    if zoek:
        like = f"%{zoek}%"
        query = query.filter(
            (Trip.startlocatie.ilike(like)) |
            (Trip.eindlocatie.ilike(like)) |
            (Trip.omschrijving.ilike(like)) |
            (Trip.odoo_partner_naam.ilike(like))
        )
    if type_filter in ("zakelijk", "prive"):
        query = query.filter_by(type=type_filter)
    if voertuig_filter:
        query = query.filter_by(vehicle_id=voertuig_filter)
    if jaar_filter:
        query = query.filter(extract("year", Trip.datum) == jaar_filter)
    if maand_filter:
        query = query.filter(extract("month", Trip.datum) == maand_filter)

    paginering = query.paginate(page=pagina, per_page=20, error_out=False)
    voertuigen = Vehicle.query.filter_by(actief=True).all()
    jaren = [r[0] for r in db.session.query(extract("year", Trip.datum)).distinct().order_by(extract("year", Trip.datum).desc()).all()]
    maand_namen = ["Jan", "Feb", "Mrt", "Apr", "Mei", "Jun", "Jul", "Aug", "Sep", "Okt", "Nov", "Dec"]

    return render_template(
        "ritten_lijst.html",
        ritten=paginering,
        zoek=zoek,
        type_filter=type_filter,
        voertuig_filter=voertuig_filter,
        jaar_filter=jaar_filter,
        maand_filter=maand_filter,
        voertuigen=voertuigen,
        jaren=jaren,
        maand_namen=maand_namen,
    )


# ── Rit toevoegen / bewerken ───────────────────────────────────────────────

@app.route("/rit/nieuw", methods=["GET", "POST"])
@app.route("/rit/<int:rit_id>/bewerk", methods=["GET", "POST"])
def rit_formulier(rit_id=None):
    rit = Trip.query.get_or_404(rit_id) if rit_id else Trip()
    default_driver = Driver.query.filter_by(naam="Pim Lavaleije").first()
    default_vehicle = Vehicle.query.filter_by(kenteken="S-581-TK").first()
    today = date.today().strftime("%Y-%m-%d")

    if request.method == "POST":
        rit.datum = datetime.strptime(request.form["datum"], "%Y-%m-%d").date()
        rit.driver_id = default_driver.id if default_driver else request.form.get("driver_id", type=int)
        rit.vehicle_id = default_vehicle.id if default_vehicle else request.form.get("vehicle_id", type=int)
        rit.startlocatie = request.form.get("startlocatie", "Thuis").strip() or "Thuis"
        rit.eindlocatie = request.form.get("eindlocatie", "").strip() or "Onbekend"
        rit.type = "zakelijk"
        rit.odoo_partner_id = request.form.get("odoo_partner_id") or None
        rit.odoo_partner_naam = request.form.get("odoo_partner_naam") or None
        rit.odoo_project_id = None
        rit.odoo_project_naam = None
        rit.omschrijving = rit.odoo_partner_naam or rit.eindlocatie or "Rit"
        rit.notitie = request.form.get("notitie", "").strip() or None

        km = Decimal(request.form["kilometers"])

        if not rit_id:
            max_eind = db.session.query(func.max(Trip.eindstand_km)).filter(
                Trip.vehicle_id == rit.vehicle_id
            ).scalar()
            begin = Decimal(str(max_eind)) if max_eind is not None else Decimal("0")
            rit.beginstand_km = begin
            rit.eindstand_km = begin + km

            fouten = rit.validate()
            if fouten:
                for f in fouten:
                    flash(f, "danger")
                return render_template("rit_formulier.html", rit=rit,
                                       default_driver=default_driver, default_vehicle=default_vehicle, today=today)

            db.session.add(rit)
            db.session.flush()

            if request.form.get("heen_en_terug"):
                rit2 = Trip(
                    datum=rit.datum,
                    driver_id=rit.driver_id,
                    vehicle_id=rit.vehicle_id,
                    startlocatie=rit.eindlocatie,
                    eindlocatie=rit.startlocatie,
                    beginstand_km=rit.eindstand_km,
                    eindstand_km=rit.eindstand_km + km,
                    type="zakelijk",
                    odoo_partner_id=rit.odoo_partner_id,
                    odoo_partner_naam=rit.odoo_partner_naam,
                    odoo_project_id=None,
                    odoo_project_naam=None,
                    omschrijving=rit.omschrijving,
                    notitie=rit.notitie,
                )
                db.session.add(rit2)
        else:
            rit.eindstand_km = rit.beginstand_km + km
            fouten = rit.validate()
            if fouten:
                for f in fouten:
                    flash(f, "danger")
                return render_template("rit_formulier.html", rit=rit,
                                       default_driver=default_driver, default_vehicle=default_vehicle, today=today)

        db.session.commit()
        flash("Rit opgeslagen.", "success")
        return redirect(url_for("ritten_lijst"))

    return render_template("rit_formulier.html", rit=rit,
                           default_driver=default_driver, default_vehicle=default_vehicle, today=today)


# ── Rit verwijderen ────────────────────────────────────────────────────────

@app.route("/rit/<int:rit_id>/verwijder", methods=["POST"])
def rit_verwijder(rit_id):
    rit = Trip.query.get_or_404(rit_id)
    db.session.delete(rit)
    db.session.commit()
    flash("Rit verwijderd.", "info")
    return redirect(url_for("ritten_lijst"))


# ── Odoo autocomplete API ──────────────────────────────────────────────────

@app.route("/api/odoo/partners")
def api_partners():
    zoek = request.args.get("q", "")
    return jsonify(odoo.get_partners(zoek))


@app.route("/api/odoo/projects")
def api_projects():
    zoek = request.args.get("q", "")
    return jsonify(odoo.get_projects(zoek))


@app.route("/api/odoo/status")
def api_odoo_status():
    import os
    return jsonify({
        "ODOO_URL": os.getenv("ODOO_URL", ""),
        "ODOO_DB": os.getenv("ODOO_DB", ""),
        "ODOO_USERNAME": os.getenv("ODOO_USERNAME", ""),
        "ODOO_PASSWORD_set": bool(os.getenv("ODOO_PASSWORD", "")),
        "configured": odoo.configured,
    })


# ── Export ─────────────────────────────────────────────────────────────────

@app.route("/export/excel")
def export_excel():
    jaar = request.args.get("jaar", date.today().year, type=int)
    maand = request.args.get("maand", 0, type=int)

    query = Trip.query.order_by(Trip.datum, Trip.beginstand_km)
    query = query.filter(extract("year", Trip.datum) == jaar)
    if maand:
        query = query.filter(extract("month", Trip.datum) == maand)
    ritten = query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ritten"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3864")
    headers = [
        "Datum", "Bestuurder", "Voertuig", "Startlocatie", "Eindlocatie",
        "Beginstand km", "Eindstand km", "Kilometers", "Type",
        "Klant", "Project", "Omschrijving", "Notitie",
    ]
    for col, h in enumerate(headers, 1):
        cel = ws.cell(row=1, column=col, value=h)
        cel.font = header_font
        cel.fill = header_fill
        cel.alignment = Alignment(horizontal="center")

    for rij, rit in enumerate(ritten, 2):
        ws.append([
            rit.datum,
            rit.driver.naam,
            rit.vehicle.kenteken,
            rit.startlocatie,
            rit.eindlocatie,
            float(rit.beginstand_km),
            float(rit.eindstand_km),
            rit.kilometers,
            rit.type,
            rit.odoo_partner_naam or "",
            rit.odoo_project_naam or "",
            rit.omschrijving,
            rit.notitie or "",
        ])

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["L"].width = 40

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    bestandsnaam = f"ritten_{jaar}{'_' + str(maand).zfill(2) if maand else ''}.xlsx"
    return send_file(output, as_attachment=True, download_name=bestandsnaam,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Import vanuit Excel ────────────────────────────────────────────────────

@app.route("/import", methods=["GET", "POST"])
def import_excel():
    if request.method == "POST":
        bestand = request.files.get("bestand")
        if not bestand or not bestand.filename.endswith((".xlsx", ".xls")):
            flash("Kies een geldig Excel-bestand (.xlsx).", "danger")
            return redirect(request.url)

        try:
            wb = load_workbook(bestand, data_only=True, read_only=True)
            ws = wb.active
            geimporteerd = 0
            fouten = []

            for rij_nr, rij in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not rij[0]:
                    continue
                try:
                    datum = rij[0] if isinstance(rij[0], date) else datetime.strptime(str(rij[0]), "%Y-%m-%d").date()
                    bestuurder_naam = str(rij[1]).strip() if rij[1] else "Onbekend"
                    kenteken = str(rij[2]).strip().upper() if rij[2] else "ONBEKEND"

                    bestuurder = Driver.query.filter_by(naam=bestuurder_naam).first()
                    if not bestuurder:
                        bestuurder = Driver(naam=bestuurder_naam)
                        db.session.add(bestuurder)
                        db.session.flush()

                    voertuig = Vehicle.query.filter_by(kenteken=kenteken).first()
                    if not voertuig:
                        voertuig = Vehicle(kenteken=kenteken)
                        db.session.add(voertuig)
                        db.session.flush()

                    type_val = str(rij[8]).lower().strip() if rij[8] else "zakelijk"
                    if type_val not in ("zakelijk", "prive", "privé"):
                        type_val = "zakelijk"
                    type_val = "prive" if "priv" in type_val else "zakelijk"

                    rit = Trip(
                        datum=datum,
                        driver_id=bestuurder.id,
                        vehicle_id=voertuig.id,
                        startlocatie=str(rij[3] or "").strip() or "Onbekend",
                        eindlocatie=str(rij[4] or "").strip() or "Onbekend",
                        beginstand_km=Decimal(str(rij[5] or 0)),
                        eindstand_km=Decimal(str(rij[6] or 0)),
                        type=type_val,
                        odoo_partner_naam=str(rij[9] or "").strip() or None,
                        odoo_project_naam=str(rij[10] or "").strip() or None,
                        omschrijving=str(rij[11] or "").strip() or "Geïmporteerd",
                        notitie=str(rij[12] or "").strip() or None,
                    )
                    validatie = rit.validate()
                    if validatie:
                        fouten.append(f"Rij {rij_nr}: {'; '.join(validatie)}")
                        continue

                    db.session.add(rit)
                    geimporteerd += 1
                except Exception as e:
                    fouten.append(f"Rij {rij_nr}: {e}")

            db.session.commit()
            flash(f"{geimporteerd} ritten geïmporteerd.", "success")
            if fouten:
                flash(f"Overgeslagen rijen: {'; '.join(fouten[:5])}", "warning")
        except Exception as e:
            flash(f"Fout bij inlezen bestand: {e}", "danger")

        return redirect(url_for("ritten_lijst"))

    return render_template("import.html")


# ── Instellingen: voertuigen & bestuurders ─────────────────────────────────

@app.route("/instellingen")
def instellingen():
    voertuigen = Vehicle.query.order_by(Vehicle.kenteken).all()
    bestuurders = Driver.query.order_by(Driver.naam).all()
    return render_template("instellingen.html", voertuigen=voertuigen, bestuurders=bestuurders)


@app.route("/voertuig/nieuw", methods=["POST"])
def voertuig_nieuw():
    v = Vehicle(
        kenteken=request.form["kenteken"].upper().strip(),
        merk=request.form.get("merk", "").strip() or None,
        model=request.form.get("model", "").strip() or None,
    )
    db.session.add(v)
    db.session.commit()
    flash(f"Voertuig {v.kenteken} toegevoegd.", "success")
    return redirect(url_for("instellingen"))


@app.route("/bestuurder/nieuw", methods=["POST"])
def bestuurder_nieuw():
    b = Driver(
        naam=request.form["naam"].strip(),
        email=request.form.get("email", "").strip() or None,
    )
    db.session.add(b)
    db.session.commit()
    flash(f"Bestuurder {b.naam} toegevoegd.", "success")
    return redirect(url_for("instellingen"))


if __name__ == "__main__":
    with app.app_context():
        db.create_all()
    app.run(debug=True, host="0.0.0.0", port=5000)
